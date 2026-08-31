import torch
import os
import time
import json
from typing import List, Dict, Any
from tqdm import tqdm

from model.inference import Inference
from sae.bank import SAEBank
from data.loader import DataLoader
from config import config
from store.circuits import circuit_store


def _parse_seed_shard() -> tuple:
    """config.discovery.seed_shard "i/k" -> (i, k), validated loudly."""
    raw = str(config.discovery.seed_shard)
    try:
        i, k = (int(x) for x in raw.split("/"))
    except ValueError:
        raise ValueError(f"discovery.seed_shard must be 'i/k', got {raw!r}")
    if k < 1 or not 0 <= i < k:
        raise ValueError(f"discovery.seed_shard needs 0 <= i < k, got {raw!r}")
    return i, k
from store.latent_stats import latent_stats
from store.top_coactivation import top_coactivation
from store.logit_context import logit_ctx
from circuit.analysis import AnalysisContext, build_analyses, run_post_circuit_analyses
from eval.node_presence import evaluate_node_presence

from circuit.probe_dataset import ProbeDatasetBuilder
from circuit.discovery.base import DiscoveryMethod
from circuit.discovery.coactivation_statistical import CoactivationStatistical
from circuit.discovery.logit_attribution import LogitAttribution
from circuit.discovery.sfc_attribution_patching import SFCAttributionPatching
from circuit.discovery.neighborhood_expansion import NeighborhoodExpansion
from circuit.discovery.top_coact_attr import TopCoactAttrDiscovery
from circuit.discovery.top_coact_expansion.mlp_top_coact_sparse_expansion import MlpTopCoactSparseExpansion
from circuit.discovery.top_coact_expansion.attn_top_coact_sparse_expansion import AttnTopCoactSparseExpansion
from circuit.discovery.top_coact_expansion.resid_top_coact_sparse_expansion import ResidTopCoactSparseExpansion
from circuit.discovery.top_coact_expansion.attn_mlp_top_coact_sparse_expansion import AttnMlpTopCoactSparseExpansion
from circuit.discovery.top_coact_expansion.attn_resid_top_coact_sparse_expansion import AttnResidTopCoactSparseExpansion
from circuit.discovery.top_coact_expansion.mlp_resid_top_coact_sparse_expansion import MlpResidTopCoactSparseExpansion
from circuit.discovery.top_coact_expansion.all_top_coact_sparse_expansion import AllTopCoactSparseExpansion
from circuit.discovery.top_coact_expansion.hard_negative_coact_sparse_expansion import HardNegativeCoactSparseExpansion
from circuit.discovery.differential_activation import DifferentialActivation
from circuit.discovery.gradient_upstream import GradientUpstreamDiscovery
from circuit.discovery.layerwise_gradient_upstream import LayerwiseGradientUpstreamDiscovery
from circuit.discovery.counterfactual_gradient import CounterfactualGradientDiscovery
from circuit.discovery.ablation_gradient import AblationGradientDiscovery
from circuit.discovery.activation_gradient import ActivationGradientDiscovery
from circuit.discovery.hybrid_gradient import HybridGradientDiscovery
from circuit.discovery.circuit_tracer_baseline import CircuitTracerBaseline


METHOD_REGISTRY: Dict[str, type[DiscoveryMethod]] = {
    "coactivation_statistical": CoactivationStatistical,
    "logit_attribution": LogitAttribution,
    "sfc_attribution_patching": SFCAttributionPatching,
    "neighborhood_expansion": NeighborhoodExpansion,
    "top_coact_attr": TopCoactAttrDiscovery,
    "mlp_top_coact_sparse_expansion": MlpTopCoactSparseExpansion,
    "attn_top_coact_sparse_expansion": AttnTopCoactSparseExpansion,
    "resid_top_coact_sparse_expansion": ResidTopCoactSparseExpansion,
    "attn_mlp_top_coact_sparse_expansion": AttnMlpTopCoactSparseExpansion,
    "attn_resid_top_coact_sparse_expansion": AttnResidTopCoactSparseExpansion,
    "mlp_resid_top_coact_sparse_expansion": MlpResidTopCoactSparseExpansion,
    "all_top_coact_sparse_expansion": AllTopCoactSparseExpansion,
    "hard_negative_coact_sparse_expansion": HardNegativeCoactSparseExpansion,
    "differential_activation": DifferentialActivation,
    "gradient_upstream": GradientUpstreamDiscovery,
    "layerwise_gradient_upstream": LayerwiseGradientUpstreamDiscovery,
    "counterfactual_gradient": CounterfactualGradientDiscovery,
    "ablation_gradient": AblationGradientDiscovery,
    "activation_gradient": ActivationGradientDiscovery,
    "hybrid_gradient": HybridGradientDiscovery,
    "circuit_tracer_baseline": CircuitTracerBaseline,
}


def _build_methods(
    inference: Inference,
    bank: SAEBank,
    avg_acts: torch.Tensor,
    probe_builder: ProbeDatasetBuilder,
) -> List[DiscoveryMethod]:
    """
    Instantiates all discovery methods listed in config.discovery.methods.

    Supported method names:
      "coactivation_statistical"  — fast, no gradients, statistical baseline
      "logit_attribution"         — two-pass gradient method (recommended)
      "sfc_attribution_patching"  — SFC-style delta×gradient node attribution + Jacobian edges
      "neighborhood_expansion"    — two-hop co-activation neighbourhood expansion (no gradients)
      "top_coact_attr"            — legacy feature-to-feature attribution (broken cross-layer)
      "mlp_top_coact_sparse_expansion"      — MLP-only two-hop expansion + full attn/resid passthrough
      "attn_top_coact_sparse_expansion"     — attn-only two-hop expansion + full MLP/resid passthrough
      "resid_top_coact_sparse_expansion"    — resid-only two-hop expansion + full attn/MLP passthrough
      "attn_mlp_top_coact_sparse_expansion" — attn+mlp expansion + full resid passthrough
      "attn_resid_top_coact_sparse_expansion" — attn+resid expansion + full mlp passthrough
      "mlp_resid_top_coact_sparse_expansion" — mlp+resid expansion + full attn passthrough
      "all_top_coact_sparse_expansion"      — all-kinds expansion (attn/mlp/resid), no passthrough
      "hard_negative_coact_sparse_expansion" — all-kinds expansion + hard-negative inhibitor search
      "differential_activation"              — pos vs neg differential activation + causal attribution
      "gradient_upstream"                    — backwards gradient BFS with per-node context switching
      "layerwise_gradient_upstream"          — layer-by-layer sweep attributing against all upstream layers (not just direct predecessors)
      "counterfactual_gradient"              — gradient attribution on contrast sequences; neg_mode config controls source: "close" (hard negatives), "random" (uniform), or "distant" (max SAE cosine distance from posctx)
      "ablation_gradient"                    — positive-context necessity discovery; ranks active upstream latents whose ablation should suppress the seed
      "hybrid_gradient"                      — runs counterfactual + ablation gradient normally, fuses their circuits, then re-evaluates
      "circuit_tracer_baseline"              — direct-effects adjacency matrix + Neumann influence propagation (Attribution Graphs analogue)
      "cluster_contrast"                     — seed-free: clusters neg_ctx sequences, KL-gradient attribution per cluster
    """
    enabled_raw = config.discovery.methods
    if isinstance(enabled_raw, list):
        enabled: List[str] = enabled_raw
    elif isinstance(enabled_raw, tuple):
        enabled = list(enabled_raw)
    else:
        enabled = []
    if not enabled:
        # Default to both main methods if config key is missing
        enabled = ["coactivation_statistical", "logit_attribution"]

    methods: List[DiscoveryMethod] = []
    for name in enabled:
        if name == "cluster_contrast":
            # Handled separately in DiscoveryWindow.run() — not a seed-based method.
            continue

        method_cls = METHOD_REGISTRY.get(name)
        if method_cls is None:
            print(f"[DiscoveryWindow] Warning: unknown discovery method '{name}' — skipped.")
            continue

        kwargs: dict = {}
        if name == "layerwise_gradient_upstream":
            kwargs["profile_first_node"] = config.discovery.layerwise_gradient_upstream.profile_first_node
        methods.append(method_cls(inference, bank, avg_acts, probe_builder, **kwargs))

    return methods


class DiscoveryWindow:
    """
    Orchestrates circuit discovery for a list of seed candidates.

    Runs all enabled discovery methods for every seed and stores every circuit that
    passes the faithfulness threshold, tagged with its source method in metadata.
    Multiple methods may find different circuits from the same seed.
    """

    def __init__(
        self,
        inference: Inference,
        bank: SAEBank,
        loader: DataLoader,
        output_dir: str = "outputs/circuits",
    ):
        self.inference = inference
        self.bank = bank
        self.loader = loader
        self.output_dir = output_dir
        os.makedirs(output_dir, exist_ok=True)

        # Zero-ablation baseline: set all non-circuit latents to 0 during patching.
        # Using latent_stats.mean_seq (conditional mean over sequences where each
        # latent fires) produced baselines too close to the original on positive
        # contexts — the same sequences where the seed fires — collapsing the
        # faithfulness denominator.  Zero ablation is an unambiguous counterfactual:
        # "what would the model do if these SAE features contributed nothing?"
        n_components = bank.n_layer * len(bank.kinds)
        self.avg_acts = torch.zeros(
            (n_components, bank.d_sae), dtype=torch.float32, device=bank.device
        )
        self.probe_builder = ProbeDatasetBuilder(inference, bank, loader)
        self.methods = _build_methods(inference, bank, self.avg_acts, self.probe_builder)

        self._analysis_context = AnalysisContext(
            n_kinds=len(bank.kinds),
            d_sae=bank.d_sae,
            kinds=bank.kinds,
            top_coactivation=top_coactivation,
            latent_stats=latent_stats,
            logit_ctx=logit_ctx,
        )
        self._analyses = build_analyses(self._analysis_context)

        enabled_raw = config.discovery.methods
        enabled_list: List[str] = list(enabled_raw) if isinstance(enabled_raw, (list, tuple)) else []
        self._run_cluster_contrast = "cluster_contrast" in enabled_list

        method_names = [type(m).__name__ for m in self.methods]
        if self._run_cluster_contrast:
            method_names.append("ClusterContrastDiscovery")
        print(f"[DiscoveryWindow] Active methods: {method_names}")
        if self._analyses:
            print(f"[DiscoveryWindow] Post-circuit analyses: {[type(a).__name__ for a in self._analyses]}")

    def run(self, candidates: List[Dict[str, Any]], save_interval: int = 10):
        """Runs all discovery methods for each seed candidate."""
        shard_i, shard_k = _parse_seed_shard()
        if shard_k > 1:
            candidates = [c for j, c in enumerate(candidates)
                          if j % shard_k == shard_i]
            print(f"--- seed_shard {shard_i}/{shard_k}: "
                  f"{len(candidates)} candidates this shard ---")
        # Resume: reload this shard's own store and skip
        # (seed, method) pairs already discovered. A seed whose method ran
        # but was REJECTED is not in the store and will re-run — rejects
        # are re-decided on resume, accepted circuits never re-fit.
        _done: set = set()
        if os.path.exists(self._store_path()):
            circuit_store.load(self._store_path())
            for _c in circuit_store.circuits.values():
                _md = _c.metadata
                _done.add((_md.get("seed_comp"), _md.get("seed_latent"),
                           _md.get("discovery_method")))
            if _done:
                print(f"--- resume: {len(_done)} (seed, method) pairs "
                      f"already in {self._store_path()} ---")
        print(f"--- Starting Discovery Window: {len(candidates)} candidates × {len(self.methods)} method(s) ---")

        # Build (comp_idx, latent_idx) → candidate lookup for seed_criteria annotation
        _cand_index: Dict[Any, Dict[str, Any]] = {
            (c["comp_idx"], c["latent_idx"]): c for c in candidates
        }

        discovered_count = 0
        task_metrics: List[Dict[str, Any]] = []
        from observability.tracking import obs

        pbar = tqdm(candidates, desc="Discovering Circuits")
        for cand in pbar:
            comp_idx = cand["comp_idx"]
            latent_idx = cand["latent_idx"]
            allowed_methods = set(cand.get("methods") or [])

            for method in self.methods:
                method_name = self._method_name(method)
                if allowed_methods and method_name not in allowed_methods:
                    continue
                if (comp_idx, latent_idx, method_name) in _done:
                    continue
                m_t0 = time.perf_counter()
                forwards_before = obs.forward_passes
                circuit = method.discover(comp_idx, latent_idx)
                _m_dt = time.perf_counter() - m_t0
                task_metrics.append(
                    {
                        "task_key": f"{cand.get('candidate_index', '?')}:{method_name}",
                        "candidate_index": cand.get("candidate_index"),
                        "comp_idx": comp_idx,
                        "latent_idx": latent_idx,
                        "method": method_name,
                        "duration_s": _m_dt,
                        "forward_pass_count": obs.forward_passes - forwards_before,
                        "accepted_circuit_count": 1 if circuit else 0,
                        "peak_cuda_memory_bytes": self._peak_cuda_memory_bytes(),
                    }
                )
                
                # If a method is slow (>1s), log its duration to the console for observability
                # if m_dt > 1.0:
                #     from observability.tracking import obs
                #     pbar.write(f"  - {method.method_name}: {m_dt:.2f}s ({obs.attempt_forward_passes} forwards)")
                
                if circuit:
                    discovered_count += 1
                    # Attach per-criterion selection scores from candidate selection
                    matched = _cand_index.get((comp_idx, latent_idx))
                    if matched and "criteria_scores" in matched:
                        circuit.metadata["seed_criteria"] = dict(matched["criteria_scores"])
                    if matched:
                        self._attach_candidate_metadata(circuit, matched)
                    circuit_store.add_circuit(circuit)
                    run_post_circuit_analyses(circuit, self._analysis_context, self._analyses)
                    self._run_node_presence_eval(circuit)
                    self._consolidate_evals(circuit)
                    pbar.set_postfix({"found": discovered_count})

                    if discovered_count % save_interval == 0:
                        self.save_store()

        self.save_store()

        cluster_count = 0
        if self._run_cluster_contrast:
            from pipeline.cluster_discovery import run_cluster_contrast_discovery
            cluster_t0 = time.perf_counter()
            forwards_before = obs.forward_passes
            cluster_circuits = run_cluster_contrast_discovery(self.inference, self.bank, self.loader)
            for c in cluster_circuits:
                circuit_store.add_circuit(c)
            cluster_count = len(cluster_circuits)
            task_metrics.append(
                {
                    "task_key": "seed_free:cluster_contrast",
                    "candidate_index": None,
                    "comp_idx": None,
                    "latent_idx": None,
                    "method": "cluster_contrast",
                    "duration_s": time.perf_counter() - cluster_t0,
                    "forward_pass_count": obs.forward_passes - forwards_before,
                    "accepted_circuit_count": cluster_count,
                    "peak_cuda_memory_bytes": self._peak_cuda_memory_bytes(),
                }
            )
            if cluster_circuits:
                self.save_store()

        cluster_suffix = f"  |  {cluster_count} cluster circuits" if cluster_count else ""
        print(f"Discovery Window complete. Found {discovered_count} faithful circuits{cluster_suffix}.")
        print(f"Total Forward Passes: {obs.forward_passes}")
        print(f"Total Forward Time: {obs.total_forward_time:.2f} s")
        if obs.forward_passes > 0:
            print(f"Average Forward Duration: {obs.total_forward_time / obs.forward_passes * 1000:.1f} ms")
        print("")
        self._print_summary_table()
        self._print_eval_stats_table()
        return task_metrics

    @staticmethod
    def _method_name(method: DiscoveryMethod) -> str:
        return str(getattr(method, "method_name", type(method).__name__))

    @staticmethod
    def _peak_cuda_memory_bytes() -> int | None:
        if not torch.cuda.is_available():
            return None
        try:
            return int(torch.cuda.max_memory_allocated())
        except Exception:
            return None

    def _run_node_presence_eval(self, circuit: Any) -> None:
        """
        Runs the posctx node-presence evaluation for CounterfactualGradientDiscovery
        circuits and merges results into circuit.metadata.
        """
        _CF_METHODS = {"counterfactual_gradient", "counterfactual_gradient_random"}
        if circuit.metadata.get("discovery_method") not in _CF_METHODS:
            return
        if not config.discovery.counterfactual_gradient.node_presence_eval:
            return

        seed_comp = circuit.metadata.get("seed_comp")
        seed_latent = circuit.metadata.get("seed_latent")
        if seed_comp is None or seed_latent is None:
            return

        from store.context import top_ctx, mid_ctx
        if not top_ctx._allocated:
            print("  [NodePresence] skipped — top_ctx not loaded")
            return
        pos_tokens = self.probe_builder.build_pos_tokens(
            seed_comp, seed_latent, top_ctx, mid_ctx,
            n_pos=config.discovery.probe_batch_size,
        )
        if pos_tokens.shape[0] == 0:
            print(f"  [NodePresence] skipped — no posctx tokens for seed {seed_comp}/{seed_latent}")
            return

        result = evaluate_node_presence(self.inference, self.bank, circuit, pos_tokens)
        circuit.metadata.setdefault("evals", {})["node_presence"] = result

    def _consolidate_evals(self, circuit: Any) -> None:
        """
        Moves top-level eval scores into the nested ``circuit.metadata["evals"]``
        dict so all evaluation results live in one place in the saved JSON.

        Handles both standard-method scores (faithfulness / sufficiency /
        completeness) and CF-gradient-specific scores (counterfactual_faithfulness /
        posctx_suppression_score).  The original top-level keys are removed after
        migration so they are not duplicated.
        """
        _EVAL_KEYS = (
            "faithfulness",
            "sufficiency",
            "completeness",
            "counterfactual_faithfulness",
            "posctx_suppression_score",
            "ablation_suppression_score",
        )
        evals = circuit.metadata.setdefault("evals", {})
        for key in _EVAL_KEYS:
            if key in circuit.metadata:
                evals[key] = circuit.metadata.pop(key)

    def _print_eval_stats_table(self):
        """Prints mean and variance for every eval field across all discovered circuits."""
        from collections import defaultdict
        from rich.console import Console
        from rich.table import Table
        from rich import box

        buckets: Dict[str, List[float]] = defaultdict(list)
        for c in circuit_store.circuits.values():
            evals = c.metadata.get("evals", {})
            for key, val in evals.items():
                if isinstance(val, dict):
                    for sub_key, sub_val in val.items():
                        if isinstance(sub_val, (int, float)):
                            buckets[f"{key}.{sub_key}"].append(float(sub_val))
                elif isinstance(val, (int, float)):
                    buckets[key].append(float(val))

        if not buckets:
            return

        table = Table(
            title="Eval Stats",
            box=box.ROUNDED,
            show_lines=False,
            header_style="bold cyan",
            title_style="bold white",
        )
        table.add_column("Eval",     style="magenta", no_wrap=True)
        table.add_column("Mean",     justify="right", style="green")
        table.add_column("Variance", justify="right", style="yellow")
        table.add_column("N",        justify="right", style="dim")

        for label in sorted(buckets):
            vals = buckets[label]
            n = len(vals)
            mean = sum(vals) / n
            variance = sum((v - mean) ** 2 for v in vals) / n if n > 1 else 0.0
            table.add_row(label, f"{mean:.4f}", f"{variance:.4f}", str(n))

        Console().print(table)

    def _print_summary_table(self):
        """Prints a Rich-formatted table of all discovered circuits sorted by faithfulness."""
        from rich.console import Console
        from rich.table import Table
        from rich import box

        circuits = list(circuit_store.circuits.values())
        if not circuits:
            return

        rows = []
        for c in circuits:
            m = c.metadata
            is_cluster = m.get("discovery_method") == "cluster_contrast"

            if is_cluster:
                cev       = m.get("evals", m)   # cluster circuits may not have been consolidated
                kl_val    = m.get("kl_loss")
                faith_val = cev.get("faithfulness")
                spec_val  = cev.get("specificity")
                kl_str    = f"{kl_val:.4f}"    if kl_val    is not None else "?"
                faith_str = f"{faith_val:.4f}" if faith_val is not None else "-"
                spec_str  = f"{spec_val:.4f}"  if spec_val  is not None else "-"
                misc_parts = [
                    f"acts={m.get('n_activators', '?')}",
                    f"inh={m.get('n_inhibitors', '?')}",
                    f"KL={kl_str}",
                    f"Faith={faith_str}",
                    f"Spec={spec_str}",
                ]
                rows.append({
                    "method":      "cluster_contrast",
                    "seed_comp":   str(m.get("cluster_id",   "?")),
                    "seed_latent": str(m.get("cluster_size", "?")),
                    "nodes":       len(c.nodes),
                    "edges":       len(c.edges),
                    "faith":       float("nan"),
                    "suff":        float("nan"),
                    "comp":        float("nan"),
                    "misc":        "  ".join(misc_parts),
                    "is_cluster":  True,
                })
            else:
                ev = m.get("evals", {})
                misc_parts = []
                if "counterfactual_faithfulness" in ev:
                    misc_parts.append(f"CF Faith: {ev['counterfactual_faithfulness']:.4f}")
                if "posctx_suppression_score" in ev:
                    misc_parts.append(f"CF Sup: {ev['posctx_suppression_score']:.4f}")
                np_ = ev.get("node_presence", {})
                if "node_presence_pct_activators" in np_:
                    misc_parts.append(f"NodPres%={np_['node_presence_pct_activators']:.1f}")
                if "node_absence_pct_inhibitors" in np_:
                    misc_parts.append(f"InhAbs%={np_['node_absence_pct_inhibitors']:.1f}")
                if "posctx_circuit_sufficiency" in np_:
                    misc_parts.append(f"CircSuff={np_['posctx_circuit_sufficiency']:.4f}")
                pa = m.get("post_analysis", {})
                _post_fmt = {
                    "coact_overlap_pct":              lambda v: f"Coact%={v:.1f}",
                    "coact_overlap_pct_activators":   lambda v: f"CoactAct%={v:.1f}",
                    "coact_overlap_pct_inhibitors":   lambda v: f"CoactInh%={v:.1f}",
                    "layer_mean":                     lambda v: f"L̄={v:.1f}",
                    "layer_std":                      lambda v: f"Lσ={v:.1f}",
                    "edge_weight_gini":               lambda v: f"Gini={v:.2f}",
                    "activity_mean":                  lambda v: f"Act={v:.0f}",
                    "rarity_pct":                     lambda v: f"Rare%={v:.1f}",
                    "top_token_consistency_pct":      lambda v: f"TokCons%={v:.1f}",
                    "internode_coact_density_pct":    lambda v: f"CoactDen%={v:.1f}",
                }
                for key, fmt in _post_fmt.items():
                    val = pa.get(key)
                    if val is not None:
                        misc_parts.append(fmt(val))
                rows.append({
                    "method":      m.get("discovery_method", "unknown"),
                    "seed_comp":   str(m.get("seed_comp",   "?")),
                    "seed_latent": str(m.get("seed_latent", "?")),
                    "nodes":       len(c.nodes),
                    "edges":       len(c.edges),
                    "faith":       ev.get("faithfulness", float("nan")),
                    "suff":        ev.get("sufficiency",  float("nan")),
                    "comp":        ev.get("completeness", float("nan")),
                    "misc":        "  ".join(misc_parts),
                    "is_cluster":  False,
                })

        # Seed-based circuits sorted by faithfulness first, then cluster contrast appended
        seed_rows    = sorted([r for r in rows if not r["is_cluster"]],
                              key=lambda r: r["faith"], reverse=True)
        cluster_rows = [r for r in rows if r["is_cluster"]]
        rows = seed_rows + cluster_rows

        table = Table(
            title="Discovered Circuits",
            box=box.ROUNDED,
            show_lines=False,
            header_style="bold cyan",
            title_style="bold white",
        )
        table.add_column("Method",        style="magenta",  no_wrap=True)
        table.add_column("Comp/Cluster",  justify="right",  style="dim")
        table.add_column("Latent/Size",   justify="right",  style="dim")
        table.add_column("Nodes",         justify="right")
        table.add_column("Edges",         justify="right")
        table.add_column("Faith",         justify="right",  style="green")
        table.add_column("Suff",          justify="right",  style="yellow")
        table.add_column("Compl",         justify="right",  style="blue")
        table.add_column("Misc",          style="cyan")

        for r in rows:
            if r["is_cluster"]:
                table.add_row(
                    r["method"],
                    r["seed_comp"],    # cluster_id
                    r["seed_latent"],  # cluster_size
                    str(r["nodes"]),
                    str(r["edges"]),
                    "-",
                    "-",
                    "-",
                    r["misc"],
                    style="cyan dim",
                )
            else:
                faith = r["faith"]
                faith_str = f"{faith:.4f}" if not (faith != faith) else "nan"
                row_style = "" if faith >= 0.5 else "dim"
                table.add_row(
                    r["method"],
                    r["seed_comp"],
                    r["seed_latent"],
                    str(r["nodes"]),
                    str(r["edges"]),
                    faith_str,
                    f"{r['suff']:.4f}",
                    f"{r['comp']:.4f}",
                    r["misc"],
                    style=row_style,
                )

        Console().print(table)

    def _print_eval_stats_table(self):
        """Prints Rich tables of aggregate stats for evals, post-analysis, and seed criteria."""
        from rich.console import Console

        circuits = [c for c in circuit_store.circuits.values()
                    if c.metadata.get("discovery_method") != "cluster_contrast"]
        if not circuits:
            return

        console = Console()

        # ── Eval & post-analysis metrics ──────────────────────────────────────
        _EVAL_METRICS = [
            ("e.faith",        lambda m: m.get("evals", {}).get("faithfulness")),
            ("e.suff",         lambda m: m.get("evals", {}).get("sufficiency")),
            ("e.compl",        lambda m: m.get("evals", {}).get("completeness")),
            ("e.cf_faith",     lambda m: m.get("evals", {}).get("counterfactual_faithfulness")),
            ("e.cf_sup",       lambda m: m.get("evals", {}).get("posctx_suppression_score")),
            ("e.np.act_pct",   lambda m: (m.get("evals", {}).get("node_presence") or {}).get("node_presence_pct_activators")),
            ("e.np.act_rate",  lambda m: (m.get("evals", {}).get("node_presence") or {}).get("node_presence_rate_mean")),
            ("e.np.inh_abs%",  lambda m: (m.get("evals", {}).get("node_presence") or {}).get("node_absence_pct_inhibitors")),
            ("e.np.inh_rate",  lambda m: (m.get("evals", {}).get("node_presence") or {}).get("node_inhibitor_rate_mean")),
            ("e.np.circ_suff", lambda m: (m.get("evals", {}).get("node_presence") or {}).get("posctx_circuit_sufficiency")),
            ("pa.coact_pct",   lambda m: (m.get("post_analysis") or {}).get("coact_overlap_pct")),
            ("pa.coact_act",   lambda m: (m.get("post_analysis") or {}).get("coact_overlap_pct_activators")),
            ("pa.coact_inh",   lambda m: (m.get("post_analysis") or {}).get("coact_overlap_pct_inhibitors")),
            ("pa.l_mean",      lambda m: (m.get("post_analysis") or {}).get("layer_mean")),
            ("pa.gini",        lambda m: (m.get("post_analysis") or {}).get("edge_weight_gini")),
            ("pa.act_mean",    lambda m: (m.get("post_analysis") or {}).get("activity_mean")),
            ("pa.rare_pct",    lambda m: (m.get("post_analysis") or {}).get("rarity_pct")),
            ("pa.tok_cons",    lambda m: (m.get("post_analysis") or {}).get("top_token_consistency_pct")),
            ("pa.coact_den",   lambda m: (m.get("post_analysis") or {}).get("internode_coact_density_pct")),
        ]
        console.print(self._build_stats_table("Eval & Post-Analysis Stats", _EVAL_METRICS, circuits))

        # ── Seed criteria scores ──────────────────────────────────────────────
        # Collect every criterion name that appears across all circuits
        all_criteria: List[str] = sorted({
            k
            for c in circuits
            for k in c.metadata.get("seed_criteria", {}).keys()
        })
        if all_criteria:
            _SC_METRICS = [
                (f"sc.{crit}", (lambda crit: lambda m: (m.get("seed_criteria") or {}).get(crit))(crit))
                for crit in all_criteria
            ]
            console.print(self._build_stats_table("Seed Criteria Scores", _SC_METRICS, circuits))

    @staticmethod
    def _build_stats_table(title: str, metrics: list, circuits: list) -> Any:
        """Builds and returns a Rich Table of per-metric aggregate statistics."""
        from rich.table import Table
        from rich import box
        import statistics

        table = Table(
            title=title,
            box=box.ROUNDED,
            show_lines=False,
            header_style="bold cyan",
            title_style="bold white",
        )
        table.add_column("Metric",  style="magenta", no_wrap=True)
        table.add_column("N",       justify="right", style="dim")
        table.add_column("Mean",    justify="right", style="green")
        table.add_column("Std",     justify="right", style="dim")
        table.add_column("Min",     justify="right", style="red")
        table.add_column("p25",     justify="right")
        table.add_column("Median",  justify="right", style="cyan")
        table.add_column("p75",     justify="right")
        table.add_column("Max",     justify="right", style="green")

        for label, extractor in metrics:
            vals = [v for c in circuits
                    if (v := extractor(c.metadata)) is not None and v == v]
            if not vals:
                continue
            n = len(vals)
            vals_sorted = sorted(vals)
            mean   = sum(vals) / n
            std    = statistics.pstdev(vals)
            lo     = vals_sorted[0]
            hi     = vals_sorted[-1]
            p25    = vals_sorted[max(0, int(n * 0.25) - 1)]
            median = statistics.median(vals_sorted)
            p75    = vals_sorted[min(n - 1, int(n * 0.75))]

            fmt = "{:.1f}" if abs(mean) >= 10 or label.endswith("pct") or "%" in label else "{:.4f}"
            table.add_row(
                label,
                str(n),
                fmt.format(mean),
                fmt.format(std),
                fmt.format(lo),
                fmt.format(p25),
                fmt.format(median),
                fmt.format(p75),
                fmt.format(hi),
            )

        return table

    def _store_path(self) -> str:
        """Shard-aware store filename: plain for 0/1, .shard<i> otherwise
        so concurrent shard processes never clobber one file."""
        i, k = _parse_seed_shard()
        name = ("discovered_circuits.pt" if k == 1
                else "discovered_circuits.shard%d.pt" % i)
        return os.path.join(self.output_dir, name)

    def save_store(self):
        """Persists the circuit store to disk."""
        path = self._store_path()
        tmp_path = f"{path}.tmp"
        circuit_store.save(tmp_path)
        os.replace(tmp_path, path)
        self._save_summary()
        self._save_summary_xlsx()

    @staticmethod
    def _attach_candidate_metadata(circuit: Any, candidate: Dict[str, Any]) -> None:
        """Copy distributed candidate provenance without changing circuit structure."""
        copy_keys = (
            "run_id",
            "worker_id",
            "candidate_index",
            "config_hash",
            "artifact_hashes",
        )
        for key in copy_keys:
            if key in candidate:
                circuit.metadata[key] = candidate[key]
        circuit.metadata.setdefault("seed_comp", candidate.get("comp_idx"))
        circuit.metadata.setdefault("seed_latent", candidate.get("latent_idx"))

    def _save_summary(self):
        """Saves a JSON summary of all discovered circuits."""
        summary = []
        for _, circuit in circuit_store.circuits.items():
            summary.append({
                "name": circuit.name,
                "uuid": circuit.uuid,
                "nodes": len(circuit.nodes),
                "edges": len(circuit.edges),
                "metadata": {
                    k: v for k, v in circuit.metadata.items()
                    if isinstance(v, (int, float, str, bool, dict))
                },
            })

        path = os.path.join(self.output_dir, "summary.json")
        tmp_path = f"{path}.tmp"
        with open(tmp_path, "w") as f:
            json.dump(summary, f, indent=2)
        os.replace(tmp_path, path)

    @staticmethod
    def _flatten(d: Dict[str, Any], prefix: str = "") -> Dict[str, Any]:
        """Recursively flattens a nested dict using dot-separated keys."""
        out: Dict[str, Any] = {}
        for k, v in d.items():
            key = f"{prefix}.{k}" if prefix else k
            if isinstance(v, dict):
                out.update(DiscoveryWindow._flatten(v, key))
            else:
                out[key] = v
        return out

    def _save_summary_xlsx(self):
        """Saves a flat Excel summary of all discovered circuits (one row per circuit)."""
        import pandas as pd

        # Column ordering: priority columns first, then any remaining
        _PRIORITY = [
            "name", "uuid", "nodes", "edges",
            "discovery_method", "seed_comp", "seed_latent",
            "evals.faithfulness", "evals.sufficiency", "evals.completeness",
            "evals.counterfactual_faithfulness", "evals.posctx_suppression_score",
            "evals.node_presence.node_presence_pct_activators",
            "evals.node_presence.node_presence_rate_mean",
            "evals.node_presence.node_absence_pct_inhibitors",
            "evals.node_presence.node_inhibitor_rate_mean",
            "evals.node_presence.posctx_circuit_sufficiency",
            "post_analysis.coact_overlap_pct",
            "post_analysis.coact_overlap_pct_activators",
            "post_analysis.coact_overlap_pct_inhibitors",
            "post_analysis.layer_mean", "post_analysis.layer_std",
            "post_analysis.layer_min", "post_analysis.layer_max",
            "post_analysis.edge_weight_gini",
            "post_analysis.activity_mean", "post_analysis.activity_median",
            "post_analysis.rarity_pct",
            "post_analysis.top_token_consistency_pct",
            "post_analysis.internode_coact_density_pct",
            # Seed selection criteria scores
            "seed_criteria.logit_impact",
            "seed_criteria.logit_specificity",
            "seed_criteria.logit_diversity",
            "seed_criteria.last_token_activity",
            "seed_criteria.surprise",
            "seed_criteria.context_coherence",
            "seed_criteria.activation_variance",
            "seed_criteria.activation_skew",
            "seed_criteria.burstiness",
            "seed_criteria.connectivity",
            "seed_criteria.coactivation_diversity",
            "seed_criteria.cross_layer_reach",
            "seed_criteria.cross_component_breadth",
            "seed_criteria.pagerank_centrality",
            "seed_criteria.coactivation_uniqueness",
            "seed_criteria.top_ctx_saturation",
            "seed_criteria.mid_ctx_richness",
            "seed_criteria.pos_neg_contrast",
            "seed_criteria.activation_entropy",
            "seed_criteria.stratified_random",
            "seed_criteria.circuit_yield",
        ]

        # Short display names for known columns (applied to both sheets)
        _ALIASES: Dict[str, str] = {
            "discovery_method":                                     "method",
            "seed_latent":                                          "seed_lat",
            "n_activators":                                         "n_act",
            "n_inhibitors":                                         "n_inh",
            "evals.faithfulness":                                   "e.faith",
            "evals.sufficiency":                                    "e.suff",
            "evals.completeness":                                   "e.compl",
            "evals.counterfactual_faithfulness":                    "e.cf_faith",
            "evals.posctx_suppression_score":                       "e.cf_sup",
            "evals.node_presence.node_presence_pct_activators":     "e.np.act_pct",
            "evals.node_presence.node_presence_rate_mean":          "e.np.act_rate",
            "evals.node_presence.node_absence_pct_inhibitors":      "e.np.inh_abs_pct",
            "evals.node_presence.node_inhibitor_rate_mean":         "e.np.inh_rate",
            "evals.node_presence.posctx_circuit_sufficiency":       "e.np.circ_suff",
            "post_analysis.coact_overlap_pct":                      "pa.coact_pct",
            "post_analysis.coact_overlap_pct_activators":           "pa.coact_act",
            "post_analysis.coact_overlap_pct_inhibitors":           "pa.coact_inh",
            "post_analysis.layer_mean":                             "pa.l_mean",
            "post_analysis.layer_std":                              "pa.l_std",
            "post_analysis.layer_min":                              "pa.l_min",
            "post_analysis.layer_max":                              "pa.l_max",
            "post_analysis.edge_weight_gini":                       "pa.gini",
            "post_analysis.activity_mean":                          "pa.act_mean",
            "post_analysis.activity_median":                        "pa.act_med",
            "post_analysis.rarity_pct":                             "pa.rare_pct",
            "post_analysis.top_token_consistency_pct":              "pa.tok_cons",
            "post_analysis.internode_coact_density_pct":            "pa.coact_den",
            "seed_criteria.logit_impact":                           "sc.logit_imp",
            "seed_criteria.logit_specificity":                      "sc.logit_spec",
            "seed_criteria.logit_diversity":                        "sc.logit_div",
            "seed_criteria.last_token_activity":                    "sc.last_tok",
            "seed_criteria.surprise":                               "sc.surprise",
            "seed_criteria.context_coherence":                      "sc.ctx_coh",
            "seed_criteria.activation_variance":                    "sc.act_var",
            "seed_criteria.activation_skew":                        "sc.act_skew",
            "seed_criteria.burstiness":                             "sc.burst",
            "seed_criteria.connectivity":                           "sc.connect",
            "seed_criteria.coactivation_diversity":                 "sc.coact_div",
            "seed_criteria.cross_layer_reach":                      "sc.x_layer",
            "seed_criteria.cross_component_breadth":                "sc.x_comp",
            "seed_criteria.pagerank_centrality":                    "sc.pagerank",
            "seed_criteria.coactivation_uniqueness":                "sc.coact_uniq",
            "seed_criteria.top_ctx_saturation":                     "sc.top_sat",
            "seed_criteria.mid_ctx_richness":                       "sc.mid_rich",
            "seed_criteria.pos_neg_contrast":                       "sc.pos_neg",
            "seed_criteria.activation_entropy":                     "sc.act_ent",
            "seed_criteria.stratified_random":                      "sc.random",
            "seed_criteria.circuit_yield":                          "sc.cy_yield",
        }

        rows = []
        for _, circuit in circuit_store.circuits.items():
            row: Dict[str, Any] = {
                "name":  circuit.name,
                "uuid":  circuit.uuid,
                "nodes": len(circuit.nodes),
                "edges": len(circuit.edges),
            }
            flat_meta = self._flatten({
                k: v for k, v in circuit.metadata.items()
                if isinstance(v, (int, float, str, bool, dict))
            })
            row.update(flat_meta)
            rows.append(row)

        if not rows:
            return

        df = pd.DataFrame(rows)

        # Reorder: priority columns first (those that exist), then any remaining
        existing_priority = [c for c in _PRIORITY if c in df.columns]
        remaining = [c for c in df.columns if c not in set(_PRIORITY)]
        df = df[existing_priority + remaining]

        # Apply short display names
        df = df.rename(columns=_ALIASES)

        path = os.path.join(self.output_dir, "summary.xlsx")
        tmp_path = f"{path}.tmp.xlsx"
        with pd.ExcelWriter(tmp_path, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name="Circuits", index=False)
            self._write_correlation_sheet(writer, df)
        os.replace(tmp_path, path)

    @staticmethod
    def _write_correlation_sheet(writer: Any, df: Any) -> None:
        """Writes a correlation matrix of all numeric columns to a second sheet."""
        from openpyxl.styles import PatternFill, Font, Alignment
        from openpyxl.utils import get_column_letter

        numeric_df = df.select_dtypes(include="number")
        if numeric_df.shape[1] < 2:
            return

        corr = numeric_df.corr()

        ws = writer.book.create_sheet("Correlations")

        # Header row (column labels)
        ws.cell(row=1, column=1, value="").font = Font(bold=True)
        for col_idx, col_name in enumerate(corr.columns, start=2):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", text_rotation=45, wrap_text=False)

        # Data rows
        for row_idx, row_name in enumerate(corr.index, start=2):
            ws.cell(row=row_idx, column=1, value=row_name).font = Font(bold=True)
            for col_idx, col_name in enumerate(corr.columns, start=2):
                val = corr.loc[row_name, col_name]
                cell = ws.cell(row=row_idx, column=col_idx, value=round(float(val), 4) if val == val else None)
                cell.alignment = Alignment(horizontal="center")

                # Colour-code by correlation strength: green = positive, red = negative
                if val == val and row_name != col_name:
                    intensity = min(int(abs(val) * 200), 200)
                    if val > 0:
                        fill_colour = f"00{intensity:02X}00"
                        font_colour = "FFFFFF" if intensity > 100 else "000000"
                    else:
                        fill_colour = f"{intensity:02X}0000"
                        font_colour = "FFFFFF" if intensity > 100 else "000000"
                    cell.fill = PatternFill("solid", fgColor=fill_colour)
                    cell.font = Font(color=font_colour)
                elif row_name == col_name:
                    cell.fill = PatternFill("solid", fgColor="CCCCCC")

        # Auto-size row-label column
        max_len = max((len(str(n)) for n in corr.index), default=10)
        ws.column_dimensions["A"].width = max(max_len + 2, 12)

        # Fixed narrow width for value columns
        for col_idx in range(2, len(corr.columns) + 2):
            ws.column_dimensions[get_column_letter(col_idx)].width = 7

        # Freeze the header row and label column
        ws.freeze_panes = "B2"


def run_discovery_window(
    inference: Inference,
    bank: SAEBank,
    loader: DataLoader,
    candidates_path: str = "outputs/candidates.pt",
):
    """Entry point to run a discovery window from saved candidates."""
    if not os.path.exists(candidates_path):
        print(f"Error: Candidates file not found at {candidates_path}. Run candidate selection first.")
        return

    if not latent_stats._allocated:
        latent_stats.load("outputs/latent_stats.pt")
    if not top_coactivation._allocated:
        top_coactivation.load("outputs/top_coactivation.pt")
    if not logit_ctx._allocated:
        logit_ctx.load("outputs/logit_ctx.pt")

    from store.context import neg_ctx
    if not neg_ctx._allocated:
        neg_ctx.load("outputs/neg_ctx.pt")

    candidates = torch.load(candidates_path, weights_only=False)

    window = DiscoveryWindow(inference, bank, loader)
    window.run(candidates)
